import os
import re
import zipfile
from io import BytesIO
from datetime import datetime

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload
from openpyxl import Workbook
from openpyxl.styles import Font, numbers

from backend.database import get_db
from ..models import (
    User, Ledger, Transaction, JournalEntry, Account,
    BankAccount, Budget, BudgetLine, Receipt
)
from ..auth import get_current_active_user, get_current_ledger

router = APIRouter(prefix="/exports", tags=["exports"])


def _safe_filename(name: str) -> str:
    """Sanitize a string for use in filenames."""
    return re.sub(r'[^\w\-.]', '_', name)


def _auto_width(ws):
    """Set column widths based on content."""
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 40)


@router.get("/excel")
def export_excel(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
    current_ledger: Ledger = Depends(get_current_ledger)
):
    """Export all ledger data as an Excel file."""
    wb = Workbook()
    bold = Font(bold=True)

    # --- Sheet 1: Transaksjoner ---
    ws = wb.active
    ws.title = "Transaksjoner"
    headers = ["ID", "Dato", "Beskrivelse", "Referanse", "Status", "Kilde",
               "Kontonummer", "Kontonavn", "Debet", "Kredit", "Opprettet"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = bold

    transactions = db.query(Transaction).filter(
        Transaction.ledger_id == current_ledger.id
    ).options(
        joinedload(Transaction.journal_entries).joinedload(JournalEntry.account)
    ).order_by(Transaction.transaction_date).all()

    for txn in transactions:
        for entry in txn.journal_entries:
            ws.append([
                txn.id,
                txn.transaction_date,
                txn.description,
                txn.reference,
                txn.status.value if txn.status else "",
                txn.source.value if txn.source else "",
                entry.account.account_number if entry.account else "",
                entry.account.account_name if entry.account else "",
                float(entry.debit) if entry.debit else 0,
                float(entry.credit) if entry.credit else 0,
                txn.created_at.strftime("%Y-%m-%d %H:%M") if txn.created_at else "",
            ])

    # Format columns
    for row in ws.iter_rows(min_row=2, min_col=2, max_col=2):
        for cell in row:
            cell.number_format = 'YYYY-MM-DD'
    for row in ws.iter_rows(min_row=2, min_col=9, max_col=10):
        for cell in row:
            cell.number_format = '#,##0.00'
    _auto_width(ws)

    # --- Sheet 2: Kontoplan ---
    ws2 = wb.create_sheet("Kontoplan")
    headers2 = ["Kontonummer", "Kontonavn", "Kontotype", "Aktiv"]
    ws2.append(headers2)
    for cell in ws2[1]:
        cell.font = bold

    accounts = db.query(Account).filter(
        Account.ledger_id == current_ledger.id
    ).order_by(Account.account_number).all()

    for acc in accounts:
        ws2.append([
            acc.account_number,
            acc.account_name,
            acc.account_type.value if acc.account_type else "",
            "Ja" if acc.is_active else "Nei",
        ])
    _auto_width(ws2)

    # --- Sheet 3: Bankkontoer ---
    ws3 = wb.create_sheet("Bankkontoer")
    headers3 = ["Navn", "Type", "Kontonummer", "Saldo", "Regnskapskonto"]
    ws3.append(headers3)
    for cell in ws3[1]:
        cell.font = bold

    bank_accounts = db.query(BankAccount).filter(
        BankAccount.ledger_id == current_ledger.id
    ).options(joinedload(BankAccount.account)).all()

    for ba in bank_accounts:
        linked = f"{ba.account.account_number} {ba.account.account_name}" if ba.account else ""
        ws3.append([
            ba.name,
            ba.account_type.value if ba.account_type else "",
            ba.account_number or "",
            float(ba.balance) if ba.balance else 0,
            linked,
        ])
    for row in ws3.iter_rows(min_row=2, min_col=4, max_col=4):
        for cell in row:
            cell.number_format = '#,##0.00'
    _auto_width(ws3)

    # --- Sheet 4: Budsjett ---
    ws4 = wb.create_sheet("Budsjett")
    headers4 = ["Budsjett", "År", "Kontonummer", "Kontonavn", "Periode", "Beløp"]
    ws4.append(headers4)
    for cell in ws4[1]:
        cell.font = bold

    budgets = db.query(Budget).filter(
        Budget.ledger_id == current_ledger.id
    ).options(
        joinedload(Budget.lines).joinedload(BudgetLine.account)
    ).all()

    for budget in budgets:
        for line in budget.lines:
            ws4.append([
                budget.name,
                budget.year,
                line.account.account_number if line.account else "",
                line.account.account_name if line.account else "",
                line.period,
                float(line.amount) if line.amount else 0,
            ])
    for row in ws4.iter_rows(min_row=2, min_col=6, max_col=6):
        for cell in row:
            cell.number_format = '#,##0.00'
    _auto_width(ws4)

    # Write to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    safe_name = _safe_filename(current_ledger.name)
    filename = f"regnskap_{safe_name}_{datetime.utcnow().strftime('%Y%m%d')}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


@router.get("/receipts")
def export_receipts(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
    current_ledger: Ledger = Depends(get_current_ledger)
):
    """Export all receipt files as a ZIP archive."""
    receipts = db.query(Receipt).filter(
        Receipt.ledger_id == current_ledger.id
    ).all()

    if not receipts:
        raise HTTPException(status_code=404, detail="Ingen bilag funnet")

    output = BytesIO()
    with zipfile.ZipFile(output, 'w', zipfile.ZIP_DEFLATED) as zf:
        for receipt in receipts:
            if receipt.image_path and os.path.exists(receipt.image_path):
                arcname = f"{receipt.id}_{receipt.original_filename or os.path.basename(receipt.image_path)}"
                zf.write(receipt.image_path, arcname)

    output.seek(0)

    safe_name = _safe_filename(current_ledger.name)
    filename = f"bilag_{safe_name}_{datetime.utcnow().strftime('%Y%m%d')}.zip"

    return StreamingResponse(
        output,
        media_type="application/zip",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )
